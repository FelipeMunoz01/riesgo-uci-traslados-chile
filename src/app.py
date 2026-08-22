"""
Maqueta: Sistema de Priorización de Camas Críticas (UCI/UTI)

Estima, a partir de datos conocidos al momento de decidir un traslado
(diagnóstico, edad, comorbilidades, procedimiento previsto, tipo de ingreso,
hospital de destino), la probabilidad de que un paciente requiera cama crítica
(UCI/UTI) durante su hospitalización.

Entrenado con datos públicos GRD de Chile (2019-2024, DEIS/MINSAL).

IMPORTANTE: esta es una herramienta de apoyo a la gestión basada en patrones
históricos de codificación clínica, NO un score clínico validado de
deterioro en tiempo real (como NEWS2). No reemplaza el juicio clínico.
"""

import datetime
from pathlib import Path

import joblib
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

RAIZ = Path(__file__).resolve().parent.parent
MODELO_PATH = RAIZ / "models" / "modelo_riesgo_uci.joblib"
NORMA_MINSAL_PATH = RAIZ / "referencias" / "Norma 2018-2019 MINSAL.xlsx"
IRGRD_DICT_PATH = RAIZ / "referencias" / "Tablas maestras bases GRD 2.xlsx"

st.set_page_config(
    page_title="Priorización de Camas Críticas UCI/UTI",
    page_icon="🏥",
    layout="wide",
)

NOMBRES_LEGIBLES = {
    "cod_hospital": "Hospital",
    "sexo": "Sexo",
    "prevision": "Previsión",
    "tipo_procedencia": "Procedencia",
    "tipo_ingreso": "Tipo de ingreso",
    "diagnostico1_categoria": "Diagnóstico principal",
    "procedimiento_principal": "Procedimiento principal",
    "tiene_mcc": "Tiene MCC",
    "tiene_cc": "Tiene CC",
    "nivel_severidad_potencial": "Nivel de severidad APR potencial",
    "tiene_diabetes": "Diabetes",
    "tiene_hipertension": "Hipertensión",
    "tiene_erc": "Enfermedad renal crónica",
    "tiene_epoc": "EPOC",
    "tiene_obesidad": "Obesidad",
}
COMORBILIDADES = ["tiene_diabetes", "tiene_hipertension", "tiene_erc", "tiene_epoc", "tiene_obesidad"]


@st.cache_resource
def cargar_modelo():
    return joblib.load(MODELO_PATH)


@st.cache_resource
def cargar_norma_minsal():
    """Norma nacional MINSAL 2018-2019 por código GRD: estancia esperada, percentiles y Exitus.

    El código GRD ya asignado (no solo el diagnóstico) no se usa como feature del modelo -
    se calcula al cierre del episodio y usarlo como input sería fuga de datos. Pero como
    contexto de apoyo para un comité que ya tiene un GRD provisional, es información oficial
    valiosa: la norma nacional para ese GRD+severidad exacto.
    """
    import openpyxl

    wb_dict = openpyxl.load_workbook(IRGRD_DICT_PATH, read_only=True, data_only=True)
    ws_dict = wb_dict["IR - GRD"]
    descripciones = {}
    for cod, desc in ws_dict.iter_rows(min_row=2, values_only=True):
        if cod is not None:
            descripciones[str(cod).zfill(6)] = desc

    wb_norma = openpyxl.load_workbook(NORMA_MINSAL_PATH, read_only=True, data_only=True)
    ws_norma = wb_norma["Norma 2018-2019 MINSAL"]
    norma = {}
    for row in ws_norma.iter_rows(min_row=2, values_only=True):
        (grd, _tipo, gravedad, total_altas, _total_est, est_media, _altas_depu, _total_est_depu,
         est_media_depu, _n_out_inf, n_out_sup, exitus, p25, p50, p75, corte_inf, corte_sup,
         _peso_total, _peso_total_depu) = row
        if grd is None:
            continue
        cod = str(grd).zfill(6)
        norma[cod] = {
            "descripcion": descripciones.get(cod, "Sin descripción"),
            "gravedad": gravedad,
            "total_altas": total_altas or 0,
            "est_media": float(str(est_media).replace(",", ".")) if est_media else 0.0,
            "est_media_depurada": float(str(est_media_depu).replace(",", ".")) if est_media_depu else 0.0,
            "exitus": exitus or 0,
            "tasa_exitus": (exitus / total_altas * 100) if total_altas else 0.0,
            "p25": p25, "p50": p50, "p75": p75,
            "corte_inf": corte_inf, "corte_sup": corte_sup,
            "n_outliers_superiores": n_out_sup or 0,
        }
    return norma


def clasificar_riesgo(p, umbrales):
    if p >= umbrales["rojo"]:
        return "🔴 Alto", "#e74c3c"
    elif p >= umbrales["amarillo"]:
        return "🟡 Medio", "#f39c12"
    else:
        return "🟢 Bajo", "#2ecc71"


def gauge_riesgo(p, umbrales, color):
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=p * 100,
        number={"suffix": "%", "font": {"size": 40}},
        gauge={
            "axis": {"range": [0, 100]},
            "bar": {"color": color},
            "steps": [
                {"range": [0, umbrales["amarillo"] * 100], "color": "rgba(46,204,113,0.25)"},
                {"range": [umbrales["amarillo"] * 100, umbrales["rojo"] * 100], "color": "rgba(243,156,18,0.25)"},
                {"range": [umbrales["rojo"] * 100, 100], "color": "rgba(231,76,60,0.25)"},
            ],
        },
    ))
    fig.update_layout(height=220, margin=dict(l=20, r=20, t=10, b=10))
    return fig


def construir_input(datos: dict, artefacto: dict) -> pd.DataFrame:
    fila = {col: datos[col] for col in artefacto["features_categoricas"] + artefacto["features_numericas"]}
    df = pd.DataFrame([fila])
    for col in artefacto["features_categoricas"]:
        categorias = artefacto["opciones"][col]
        df[col] = pd.Categorical(df[col].astype(str), categories=categorias)
    return df


MAX_FILAS_CSV = 5000

# rangos aceptables por variable numérica; None = sin tope
RANGOS_NUMERICOS = {
    "edad": (0, 110),
    "mes_ingreso": (1, 12),
    "tiene_mcc": (0, 1), "tiene_cc": (0, 1),
    "nivel_severidad_potencial": (1, 3),
    "tiene_diabetes": (0, 1), "tiene_hipertension": (0, 1),
    "tiene_erc": (0, 1), "tiene_epoc": (0, 1), "tiene_obesidad": (0, 1),
}


def ejemplo_plantilla(artefacto: dict) -> dict:
    """Fila de ejemplo con valores que el modelo sí reconoce, para la plantilla."""
    fila = {}
    for col in artefacto["features_categoricas"]:
        opts = artefacto["opciones"][col]
        preferidos = {"procedimiento_principal": "SIN_PROCEDIMIENTO"}
        fila[col] = preferidos.get(col) if preferidos.get(col) in opts else opts[0]
    for col in artefacto["features_numericas"]:
        lo, hi = RANGOS_NUMERICOS.get(col, (0, 0))
        fila[col] = {"edad": 65, "mes_ingreso": 1, "nivel_severidad_potencial": 1}.get(col, lo)
    fila["_id"] = "Ejemplo - reemplazar"
    return fila


def validar_csv(bruto: pd.DataFrame, artefacto: dict):
    """Devuelve (filas_validas, lista_de_problemas).

    Sin esto la app aceptaba su propia plantilla vacía y devolvía un riesgo de
    1.85% construido enteramente con valores faltantes, sin ninguna advertencia.
    """
    cat = artefacto["features_categoricas"]
    num = artefacto["features_numericas"]
    problemas = []

    faltantes = [c for c in cat + num if c not in bruto.columns]
    if faltantes:
        return None, [f"Faltan columnas obligatorias: {', '.join(faltantes)}"]

    df = bruto.copy()
    if len(df) > MAX_FILAS_CSV:
        problemas.append(f"El archivo trae {len(df):,} filas; se procesan las primeras {MAX_FILAS_CSV:,}.")
        df = df.head(MAX_FILAS_CSV)

    por_defecto = pd.Series([f"Fila {i + 1}" for i in range(len(df))], index=df.index)
    if "_id" not in df.columns:
        df["_id"] = por_defecto
    else:
        etiquetas = df["_id"].fillna("").astype(str).str.strip()
        df["_id"] = etiquetas.where(etiquetas != "", por_defecto)

    valida = pd.Series(True, index=df.index)

    for col in num:
        convertida = pd.to_numeric(df[col], errors="coerce")
        malos = convertida.isna()
        if malos.any():
            problemas.append(f"'{col}': {malos.sum()} fila(s) con valor vacío o no numérico.")
        lo, hi = RANGOS_NUMERICOS.get(col, (None, None))
        if lo is not None:
            fuera = convertida.notna() & ((convertida < lo) | (convertida > hi))
            if fuera.any():
                problemas.append(f"'{col}': {fuera.sum()} fila(s) fuera del rango {lo}-{hi}.")
            malos = malos | fuera
        df[col] = convertida
        valida &= ~malos

    for col in cat:
        permitidos = set(artefacto["opciones"][col])
        vals = df[col].fillna("").astype(str).str.strip()
        malos = ~vals.isin(permitidos)
        if malos.any():
            ejemplos = sorted(set(vals[malos]) - {""})[:3]
            detalle = f" (ej. {', '.join(repr(e) for e in ejemplos)})" if ejemplos else ""
            problemas.append(
                f"'{col}': {malos.sum()} fila(s) con un valor que el modelo no conoce{detalle}."
            )
        df[col] = vals
        valida &= ~malos

    descartadas = int((~valida).sum())
    if descartadas == len(df):
        problemas.append(f"Ninguna de las {len(df)} filas es utilizable.")
    elif descartadas:
        problemas.append(
            f"Se descartan {descartadas} de {len(df)} filas; las otras {len(df) - descartadas} sí se pueden usar."
        )
    return df[valida].copy(), problemas


def predecir_lote(casos: list[dict], artefacto: dict, modelo) -> "pd.Series":
    """Una sola llamada a predict_proba para toda la cola.

    Antes se llamaba fila por fila dentro de un bucle de Python, lo que hacía
    que una cola grande tardara segundos y bloqueara la app.
    """
    cols = artefacto["features_categoricas"] + artefacto["features_numericas"]
    X = pd.DataFrame([{c: caso.get(c) for c in cols} for caso in casos])
    for col in artefacto["features_categoricas"]:
        X[col] = pd.Categorical(X[col].astype(str), categories=artefacto["opciones"][col])
    for col in artefacto["features_numericas"]:
        X[col] = pd.to_numeric(X[col], errors="coerce")
    return modelo.predict_proba(X)[:, 1]


def panel_explicabilidad(datos: dict, artefacto: dict):
    tasas = artefacto["tasas_historicas"]
    base = artefacto["tasa_base_2024"]
    nombres_hosp = artefacto["opciones"]["cod_hospital_nombre"]
    filas = []
    for col in artefacto["features_categoricas"]:
        valor = str(datos[col])
        tasa_valor = tasas.get(col, {}).get(valor)
        if tasa_valor is None:
            continue
        diferencia = tasa_valor - base
        valor_legible = nombres_hosp.get(valor, valor) if col == "cod_hospital" else valor
        filas.append({
            "Factor": NOMBRES_LEGIBLES.get(col, col),
            "Valor": valor_legible,
            "Tasa histórica UCI/UTI": tasa_valor,
            "vs. promedio": diferencia,
        })
    for col in COMORBILIDADES + ["tiene_mcc", "tiene_cc"]:
        valor = int(datos.get(col, 0))
        tasa_valor = tasas.get(col, {}).get(valor)
        if tasa_valor is None or valor == 0:
            continue
        diferencia = tasa_valor - base
        filas.append({
            "Factor": NOMBRES_LEGIBLES.get(col, col),
            "Valor": "Sí",
            "Tasa histórica UCI/UTI": tasa_valor,
            "vs. promedio": diferencia,
        })
    if not filas:
        return
    tabla = pd.DataFrame(filas).sort_values("vs. promedio", ascending=False, key=abs)
    tabla["Tasa histórica UCI/UTI"] = (tabla["Tasa histórica UCI/UTI"] * 100).round(1).astype(str) + "%"
    tabla["vs. promedio"] = tabla["vs. promedio"].apply(lambda x: f"{'+' if x >= 0 else ''}{x*100:.1f} pp")
    st.caption(
        "Tasa histórica de necesidad de UCI/UTI observada en los datos de entrenamiento para cada "
        f"valor elegido, comparada con el promedio general ({base*100:.1f}%). No es la explicación "
        "exacta del modelo (que combina las variables de forma no lineal), pero da una idea de qué "
        "factores empujan el riesgo hacia arriba o abajo."
    )
    st.dataframe(tabla, width="stretch", hide_index=True)


def formulario_paciente(artefacto: dict, key_prefix: str) -> dict:
    opciones = artefacto["opciones"]
    nombres_hosp = opciones["cod_hospital_nombre"]
    desc_diag = opciones["diagnostico1_descripcion"]

    col1, col2, col3 = st.columns(3)
    with col1:
        hospital = st.selectbox(
            "Hospital",
            options=opciones["cod_hospital"],
            format_func=lambda c: f"{nombres_hosp.get(c, c)} ({c})",
            key=f"{key_prefix}_hosp",
        )
        sexo = st.selectbox("Sexo", options=opciones["sexo"], key=f"{key_prefix}_sexo")
        edad = st.number_input("Edad", min_value=0, max_value=110, value=60, key=f"{key_prefix}_edad")
    with col2:
        tipo_ingreso = st.selectbox(
            "Tipo de ingreso", options=opciones["tipo_ingreso"], key=f"{key_prefix}_tingreso"
        )
        tipo_procedencia = st.selectbox(
            "Procedencia", options=opciones["tipo_procedencia"], key=f"{key_prefix}_tproc"
        )
    with col3:
        prevision = st.selectbox("Previsión", options=opciones["prevision"], key=f"{key_prefix}_prev")
        diagnostico = st.selectbox(
            "Diagnóstico principal (CIE-10)",
            options=opciones["diagnostico1_categoria"],
            format_func=lambda c: f"{c} - {desc_diag.get(c, '')}",
            key=f"{key_prefix}_diag",
        )
        mes = st.selectbox(
            "Mes de ingreso",
            options=list(range(1, 13)),
            index=datetime.date.today().month - 1,
            key=f"{key_prefix}_mes",
        )

    st.markdown("###### Procedimiento y comorbilidades")
    st.caption(
        "El procedimiento se conoce exacto si es programado, o como intención si es de "
        "urgencia. Las comorbilidades se marcan según los diagnósticos ya conocidos."
    )
    colp1, colp3 = st.columns(2)
    with colp1:
        procedimiento = st.selectbox(
            "Procedimiento principal previsto",
            options=opciones["procedimiento_principal"],
            format_func=lambda c: "Sin procedimiento" if c == "SIN_PROCEDIMIENTO" else (
                "Otro (fuera de los 200 más frecuentes)" if c == "OTRO" else c
            ),
            key=f"{key_prefix}_proc",
        )
    with colp3:
        tiene_diabetes = st.checkbox("Diabetes", key=f"{key_prefix}_diab")
        tiene_hipertension = st.checkbox("Hipertensión", key=f"{key_prefix}_hta")
        tiene_erc = st.checkbox("Enfermedad renal crónica", key=f"{key_prefix}_erc")
        tiene_epoc = st.checkbox("EPOC", key=f"{key_prefix}_epoc")
        tiene_obesidad = st.checkbox("Obesidad", key=f"{key_prefix}_obes")

    st.caption(
        "Severidad APR (según Módulo APR CC/MCC del IR-GRD chileno) — marca si algún "
        "diagnóstico secundario ya codificado corresponde a alguna de estas categorías:"
    )
    colm1, colm2 = st.columns(2)
    with colm1:
        tiene_mcc = st.checkbox(
            "Tiene MCC (complicación/comorbilidad MAYOR)", key=f"{key_prefix}_mcc",
            help="Ej: shock séptico, sepsis, insuficiencia respiratoria aguda, IRA, desnutrición moderada-severa.",
        )
    with colm2:
        tiene_cc = st.checkbox(
            "Tiene CC (complicación/comorbilidad, sin MCC)", key=f"{key_prefix}_cc",
            help="Ej: FA, EPOC exacerbado, ERC etapa 3-4, delirium, anemia activa.",
        )
    nivel_severidad_potencial = 3 if tiene_mcc else (2 if tiene_cc else 1)

    return {
        "cod_hospital": hospital,
        "sexo": sexo,
        "edad": edad,
        "mes_ingreso": mes,
        "prevision": prevision,
        "tipo_procedencia": tipo_procedencia,
        "tipo_ingreso": tipo_ingreso,
        "diagnostico1_categoria": diagnostico,
        "procedimiento_principal": procedimiento,
        "tiene_mcc": int(tiene_mcc),
        "tiene_cc": int(tiene_cc),
        "nivel_severidad_potencial": nivel_severidad_potencial,
        "tiene_diabetes": int(tiene_diabetes),
        "tiene_hipertension": int(tiene_hipertension),
        "tiene_erc": int(tiene_erc),
        "tiene_epoc": int(tiene_epoc),
        "tiene_obesidad": int(tiene_obesidad),
    }


DIVERGENCIA_ALERTA_PP = 5.0  # puntos porcentuales


def tab_caso_individual(artefacto):
    cama, clinico = artefacto["cama_critica"], artefacto["riesgo_clinico"]
    datos = formulario_paciente(artefacto, "individual")
    if st.button("Calcular riesgo", type="primary"):
        X = construir_input(datos, artefacto)
        p_cama = cama["modelo"].predict_proba(X)[0, 1]
        p_clin = clinico["modelo"].predict_proba(X)[0, 1]
        et_cama, col_cama = clasificar_riesgo(p_cama, cama["umbrales_semaforo"])
        et_clin, col_clin = clasificar_riesgo(p_clin, clinico["umbrales_semaforo"])

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### 🏥 Probabilidad de cama crítica")
            st.caption("Pregunta de **recurso**: ¿el destino tiene dónde ponerlo?")
            st.plotly_chart(gauge_riesgo(p_cama, cama["umbrales_semaforo"], col_cama),
                            width="stretch", key="g_cama")
            st.markdown(f"### {et_cama}")
            st.caption(f"Tasa histórica de referencia: {cama['tasa_base_2024']*100:.1f}%.")
        with c2:
            st.markdown("#### 🩺 Riesgo de desenlace adverso")
            st.caption("Pregunta **clínica**: ¿este traslado es seguro? (UCI/UTI o fallecimiento)")
            st.plotly_chart(gauge_riesgo(p_clin, clinico["umbrales_semaforo"], col_clin),
                            width="stretch", key="g_clin")
            st.markdown(f"### {et_clin}")
            st.caption(f"Tasa histórica de referencia: {clinico['tasa_base_2024']*100:.1f}%.")

        divergencia = (p_clin - p_cama) * 100
        if divergencia >= DIVERGENCIA_ALERTA_PP:
            st.warning(
                f"**Revisar con atención: divergencia de {divergencia:.1f} puntos.** El riesgo "
                f"clínico ({p_clin*100:.1f}%) supera bastante la probabilidad de cama crítica "
                f"({p_cama*100:.1f}%). A pacientes con este perfil históricamente se les asignó "
                f"cama crítica con menos frecuencia de lo que su desenlace sugeriría — es un patrón "
                f"típico en adultos mayores y pacientes frágiles. La decisión de traslado no "
                f"debería apoyarse solo en la disponibilidad de cama."
            )

        nivel_txt = {1: "1 — Sin CC/MCC", 2: "2 — Con CC", 3: "3 — Con MCC"}[datos["nivel_severidad_potencial"]]
        st.metric("Nivel de severidad APR potencial", nivel_txt)
        st.caption("Según reglas CC/MCC del Módulo APR (IR-GRD Chile) — misma lógica del agrupador real.")

        st.markdown("#### ¿Por qué este resultado?")
        panel_explicabilidad(datos, artefacto)

    st.divider()
    with st.expander("🏛️ Comparar con Norma Nacional MINSAL 2018-2019 (opcional, si ya hay un GRD codificado)"):
        st.caption(
            "Si el codificador ya tiene un GRD provisional asignado a partir de la evolución médica, "
            "acá se puede consultar la norma nacional oficial para ese GRD exacto (estancia esperada, "
            "percentiles, tasa de Exitus histórica nacional) como contexto adicional al score del modelo."
        )
        norma = cargar_norma_minsal()
        opciones_grd = sorted(norma.keys())
        cod_grd = st.selectbox(
            "Código GRD",
            options=opciones_grd,
            format_func=lambda c: f"{c} - {norma[c]['descripcion']}",
            index=None,
            placeholder="Buscar código o descripción...",
            key="grd_norma",
        )
        if cod_grd:
            n = norma[cod_grd]
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Tasa de Exitus nacional", f"{n['tasa_exitus']:.1f}%", help=f"{n['exitus']:,} de {n['total_altas']:,} altas")
            c2.metric("Estancia media esperada", f"{n['est_media_depurada']:.1f} días")
            c3.metric("Percentil 50 (mediana)", f"{n['p50']} días")
            c4.metric("Punto de corte outlier", f"{n['corte_sup']} días")
            st.caption(
                f"Gravedad {n['gravedad']} — percentiles de estancia: P25={n['p25']}, P50={n['p50']}, "
                f"P75={n['p75']} días. Basado en {n['total_altas']:,} altas a nivel nacional (2018-2019)."
            )


def tab_cola(artefacto):
    st.write("Agrega casos manualmente o sube un CSV con varios pacientes, y compáralos ordenados por riesgo.")
    if "cola" not in st.session_state:
        st.session_state.cola = []

    modo = st.radio("Modo de carga", ["Formulario manual", "Subir CSV"], horizontal=True, key="modo_carga")

    if modo == "Formulario manual":
        with st.form("agregar_caso"):
            datos_cola = formulario_paciente(artefacto, "cola")
            etiqueta_libre = st.text_input("Identificador del caso (opcional, ej. cama/box)", "")
            agregar = st.form_submit_button("➕ Agregar a la cola")
            if agregar:
                st.session_state.cola.append(
                    {**datos_cola, "_id": etiqueta_libre or f"Caso {len(st.session_state.cola)+1}"}
                )
    else:
        columnas_esperadas = artefacto["features_categoricas"] + artefacto["features_numericas"]
        st.caption("El CSV debe tener columnas: " + ", ".join(columnas_esperadas) + " (opcional: _id).")
        # La plantilla trae una fila de EJEMPLO con valores reales, no vacía: antes
        # se descargaba en blanco y al subirla sin tocar devolvía un riesgo de 1.85%
        # calculado sobre puros faltantes.
        plantilla = pd.DataFrame([ejemplo_plantilla(artefacto)])
        st.download_button(
            "⬇️ Descargar plantilla CSV",
            plantilla.to_csv(index=False).encode("utf-8"),
            file_name="plantilla_pacientes.csv",
            mime="text/csv",
        )
        archivo = st.file_uploader("Subir CSV de pacientes", type=["csv"])
        if archivo is not None:
            try:
                nuevo = pd.read_csv(archivo, dtype=str)
            except Exception as e:
                st.error(f"No se pudo leer el CSV: {e}")
                nuevo = None
            if nuevo is not None:
                limpio, problemas = validar_csv(nuevo, artefacto)
                if problemas:
                    st.error("El CSV tiene problemas:\n\n" + "\n".join(f"- {p}" for p in problemas))
                if limpio is not None and not limpio.empty:
                    st.success(f"{len(limpio)} filas válidas listas para agregar.")
                    st.dataframe(limpio.head(10), width="stretch", hide_index=True)
                    if st.button(f"➕ Agregar {len(limpio)} casos del CSV a la cola"):
                        st.session_state.cola.extend(limpio.to_dict(orient="records"))
                        st.rerun()

    if st.session_state.cola:
        cama, clinico = artefacto["cama_critica"], artefacto["riesgo_clinico"]
        p_cama = predecir_lote(st.session_state.cola, artefacto, cama["modelo"])
        p_clin = predecir_lote(st.session_state.cola, artefacto, clinico["modelo"])
        nombres_hosp = artefacto["opciones"]["cod_hospital_nombre"]
        filas = [
            {
                "Caso": caso["_id"],
                "Hospital": nombres_hosp.get(str(caso["cod_hospital"]), ""),
                "Edad": caso["edad"],
                "Diagnóstico": caso["diagnostico1_categoria"],
                "Procedimiento": caso["procedimiento_principal"],
                "Cama crítica": pc,
                "Riesgo clínico": pk,
                "Divergencia": (pk - pc) * 100,
                "Nivel clínico": clasificar_riesgo(pk, clinico["umbrales_semaforo"])[0],
            }
            for caso, pc, pk in zip(st.session_state.cola, p_cama, p_clin)
        ]
        # se ordena por riesgo CLÍNICO: es la pregunta que decide si el traslado es seguro
        tabla = pd.DataFrame(filas).sort_values("Riesgo clínico", ascending=False)
        n_divergentes = int((tabla["Divergencia"] >= DIVERGENCIA_ALERTA_PP).sum())
        for c in ("Cama crítica", "Riesgo clínico"):
            tabla[c] = (tabla[c] * 100).round(1).astype(str) + "%"
        tabla["Divergencia"] = tabla["Divergencia"].apply(
            lambda x: f"⚠️ +{x:.1f} pp" if x >= DIVERGENCIA_ALERTA_PP else f"{x:+.1f} pp"
        )
        st.caption("Ordenada por riesgo clínico. La columna *Divergencia* marca los casos donde el "
                   "riesgo clínico supera a la probabilidad de cama: revisar con atención.")
        st.dataframe(tabla, width="stretch", hide_index=True)
        if n_divergentes:
            st.warning(f"{n_divergentes} caso(s) con divergencia ≥ {DIVERGENCIA_ALERTA_PP:.0f} puntos.")

        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("🗑️ Vaciar cola"):
                st.session_state.cola = []
                st.rerun()
        with col_b:
            st.download_button(
                "⬇️ Descargar cola como CSV",
                tabla.to_csv(index=False).encode("utf-8"),
                file_name="cola_priorizacion.csv",
                mime="text/csv",
            )
    else:
        st.info("Aún no has agregado casos a la cola.")


def tab_sobre_el_modelo(artefacto):
    st.markdown(f"""
    Entrenado con **{len(artefacto['opciones']['cod_hospital'])} hospitales** públicos chilenos con
    unidad crítica propia (datos GRD 2019-2023) y evaluado en **2024**, un año no visto durante
    el entrenamiento. Se excluyeron los hospitales sin unidad crítica propia, porque derivan a
    sus pacientes a otro centro y eso contaminaría la etiqueta.

    **Esta herramienta usa solo el código de diagnóstico y variables administrativas de ingreso
    — no reemplaza signos vitales ni el juicio clínico.** Es un prototipo para apoyar la
    priorización, no un dispositivo médico validado.
    """)

    st.markdown("#### Los dos modelos")
    filas = []
    for clave in ("cama_critica", "riesgo_clinico"):
        b = artefacto[clave]
        mt = b.get("metricas_traslados_2024") or {}
        filas.append({
            "Modelo": b["etiqueta"],
            "ROC-AUC": f"{b['metricas_test_2024']['roc_auc']:.3f}",
            "Average Precision": f"{b['metricas_test_2024']['average_precision']:.3f}",
            "Tasa base": f"{b['tasa_base_2024']*100:.1f}%",
            "AUC (solo traslados)": f"{mt.get('roc_auc', float('nan')):.3f}",
            "AP (solo traslados)": f"{mt.get('average_precision', float('nan')):.3f}",
        })
    st.dataframe(pd.DataFrame(filas), width="stretch", hide_index=True)

    with st.expander("⚠️ Por qué hacen falta DOS puntajes y no uno", expanded=True):
        st.markdown("""
        La etiqueta original mide **uso** de cama crítica, no necesidad clínica. Y el uso está
        sesgado por el racionamiento: a los pacientes mayores se les asigna menos UCI aunque
        estén más graves, por limitación del esfuerzo terapéutico y disponibilidad de camas.

        Medido en los datos, para **neumonía (J18)**:

        | edad | uso de UCI/UTI | mortalidad |
        |---|---|---|
        | menores de 50 | 22,3 % | 3,9 % |
        | 65 a 79 | 19,2 % | 17,6 % |
        | **80 o más** | **8,6 %** | **23,3 %** |

        El grupo de 80+ usa 2,6 veces menos UCI y muere 6 veces más. Mismo patrón en sepsis
        (55 % → 26 % de uso mientras la mortalidad sube de 18 % a 50 %).

        Un modelo entrenado solo con "¿ocupó cama crítica?" diría **riesgo bajo** justo para los
        pacientes más frágiles. Por eso se agregó el segundo puntaje, entrenado sobre
        **UCI/UTI o fallecimiento**, que no sufre ese sesgo del mismo modo.

        **Cuando los dos puntajes divergen, esa divergencia es la señal más útil**: riesgo
        clínico alto con probabilidad de cama baja marca exactamente los casos que el comité
        debe mirar dos veces.

        *Advertencia:* la mortalidad intrahospitalaria tampoco es un patrón oro puro — la
        limitación del esfuerzo terapéutico también influye en morir, no solo en recibir cama.
        Reduce mucho el sesgo, no lo elimina.
        """)

    with st.expander("¿Por qué el modelo no usa la especialidad médica ni el número de procedimientos?"):
        st.markdown("""
        Porque el GRD es una base de **egresos**: la ficha se llena al alta. Una columna solo
        sirve si su *valor* se conocía al momento de decidir, no solo si existe en el archivo.

        - `especialidad_medica` es el servicio tratante consolidado al alta. "Medicina Intensiva
          Adulto" tiene 98% de tasa de UCI/UTI: es prácticamente un sinónimo de la respuesta.
        - `n_diagnosticos_secundarios` y `n_procedimientos` cuentan casillas llenadas durante
          **toda** la estadía. Un paciente sale con 12 diagnósticos porque se complicó, no
          porque llegara así.

        Usarlas subía el ROC-AUC de 0.914 a 0.955, pero era una mejora que no existe en la vida
        real: al momento de decidir el traslado esos valores todavía no están.

        Tampoco se incluye la cirugía mayor ambulatoria (opera y se va el mismo día, 1.2% de
        tasa de UCI/UTI): no es objeto de una decisión de traslado y solo inflaba la métrica.
        """)

    COLORES = {"cama_critica": "#3498db", "riesgo_clinico": "#9b59b6"}

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("#### Curva ROC (test 2024)")
        fig = go.Figure()
        for clave in ("cama_critica", "riesgo_clinico"):
            roc = artefacto[clave]["curva_roc"]
            fig.add_trace(go.Scatter(x=roc["fpr"], y=roc["tpr"], mode="lines",
                                     name=artefacto[clave]["etiqueta"],
                                     line=dict(color=COLORES[clave], width=3)))
        fig.add_trace(go.Scatter(x=[0, 1], y=[0, 1], mode="lines", name="Azar",
                                  line=dict(color="gray", dash="dash")))
        fig.update_layout(
            xaxis_title="Falsos positivos", yaxis_title="Verdaderos positivos",
            height=350, margin=dict(l=10, r=10, t=10, b=10),
            legend=dict(x=0.35, y=0.12, font=dict(size=10)),
        )
        st.plotly_chart(fig, width="stretch")
    with c2:
        st.markdown("#### Calibración (predicho vs. real)")
        fig = go.Figure()
        for clave in ("cama_critica", "riesgo_clinico"):
            cal = artefacto[clave]["curva_calibracion"]
            fig.add_trace(go.Scatter(x=cal["mean_pred"], y=cal["frac_pos"], mode="lines+markers",
                                     name=artefacto[clave]["etiqueta"],
                                     line=dict(color=COLORES[clave], width=3)))
        fig.add_trace(go.Scatter(x=[0, 1], y=[0, 1], mode="lines", name="Ideal",
                                  line=dict(color="gray", dash="dash")))
        fig.update_layout(
            xaxis_title="Probabilidad predicha", yaxis_title="Fracción real positiva",
            height=350, margin=dict(l=10, r=10, t=10, b=10),
            legend=dict(x=0.35, y=0.12, font=dict(size=10)),
        )
        st.plotly_chart(fig, width="stretch")

    st.markdown("#### Importancia de variables")
    st.caption("Cuánto cae el ROC-AUC al mezclar aleatoriamente cada variable (permutation importance).")
    fig = go.Figure()
    orden = [f for f, _ in artefacto["cama_critica"]["importancia_variables"]]
    for clave in ("cama_critica", "riesgo_clinico"):
        imp = dict(artefacto[clave]["importancia_variables"])
        fig.add_trace(go.Bar(
            x=[imp.get(f, 0) for f in orden],
            y=[NOMBRES_LEGIBLES.get(f, f) for f in orden],
            orientation="h", name=artefacto[clave]["etiqueta"], marker_color=COLORES[clave],
        ))
    fig.update_layout(height=450, margin=dict(l=10, r=10, t=10, b=10),
                      xaxis_title="Caída en ROC-AUC", barmode="group",
                      legend=dict(x=0.5, y=0.08, font=dict(size=10)))
    st.plotly_chart(fig, width="stretch")


def main():
    artefacto = cargar_modelo()

    st.title("🏥 Priorización de Camas Críticas (UCI/UTI)")
    st.caption(
        "Maqueta de apoyo a la gestión clínica — estima la probabilidad de que un paciente "
        "requiera cama crítica según patrones históricos de casos GRD similares."
    )

    tab1, tab2, tab3 = st.tabs(["🧑‍⚕️ Evaluar caso individual", "📋 Cola de priorización", "📊 Sobre el modelo"])
    with tab1:
        tab_caso_individual(artefacto)
    with tab2:
        tab_cola(artefacto)
    with tab3:
        tab_sobre_el_modelo(artefacto)


if __name__ == "__main__":
    main()
