"""
Entrena un clasificador de riesgo de necesitar cama crítica (UCI/UTI) usando
solo variables cuyo VALOR se conoce al momento de decidir el traslado.

No basta con excluir severidad/mortalidad GRD y tipo de alta: también quedaron
fuera especialidad_medica (el servicio tratante consolidado al alta; 'MEDICINA
INTENSIVA ADULTO' tiene 98% de tasa UCI, casi un sinónimo de la etiqueta) y los
contadores n_diagnosticos_secundarios / n_procedimientos (cuentan casillas
llenadas durante toda la estadía, no lo conocido al ingreso).

Split temporal: entrena con ANIOS_TRAIN, evalúa en ANIO_TEST (año no visto), que es
más realista que un split aleatorio para saber cómo se comportaría el modelo
en producción con datos futuros.

Población: se excluye la cirugía mayor ambulatoria (CMA), que opera y se va el
mismo día y nunca es objeto de una decisión de traslado a cama crítica.

Filtro de hospitales: se excluyen los hospitales sin unidad crítica propia
(tasa histórica de UCI/UTI ~0%, señal de que derivan esos pacientes a otro
centro y por lo tanto el traslado no queda registrado en sus datos -> label
contaminado) o con muy poco volumen. cod_hospital se incluye como feature
para que el modelo aprenda el patrón/capacidad propia de cada centro.
"""

import joblib
from pathlib import Path
import numpy as np
import pandas as pd
from sklearn.calibration import calibration_curve
from sklearn.ensemble import HistGradientBoostingClassifier
from sklearn.inspection import permutation_importance
from sklearn.metrics import average_precision_score, brier_score_loss, roc_auc_score, roc_curve

RAIZ = Path(__file__).resolve().parent.parent
DATA = RAIZ / "data" / "admisiones.parquet"
MODELO_OUT = RAIZ / "models" / "modelo_riesgo_uci.joblib"
TABLAS_MAESTRAS = RAIZ / "referencias" / "Tablas maestras bases GRD 2.xlsx"
CIE10_DICT = RAIZ / "referencias" / "CIE-10.xlsx"

MIN_CASOS_HOSPITAL = 5000
MIN_TASA_HOSPITAL = 0.01  # excluye hospitales sin UCI/UTI propia (derivan pacientes)

ANIO_TEST = 2024
# 2019 subregistra la etiqueta: 9.1% vs ~14% en los demás años (ingreso directo a
# servicio crítico 6.2% vs 10-12%, SERVICIOTRASLADO1 poblado 14.4% vs 20.6%).
# Medido sobre el mismo test 2024, entrenando con cada ventana:
#   2019-2023 (3.88M filas)  AUC 0.9131  AP 0.7311
#   2020-2023 (2.93M filas)  AUC 0.9146  AP 0.7358   <- por defecto
#   2022-2023 (1.56M filas)  AUC 0.9161  AP 0.7395
# Menos datos pero más recientes rinde mejor. Se usa 2020-2023 como compromiso:
# saca el año con el defecto de registro y conserva volumen para las categorías
# poco frecuentes. Acortar a [2022, 2023] da algo más de AP a costa del 47% de las
# filas, y probablemente exigiría reentrenar más seguido.
ANIOS_TRAIN = [2020, 2021, 2022, 2023]

COMORBILIDADES = ["tiene_diabetes", "tiene_hipertension", "tiene_erc", "tiene_epoc", "tiene_obesidad"]

# Solo variables cuyo VALOR se conoce al momento de decidir el traslado.
# Ver el docstring de extract_data.py para por qué especialidad_medica,
# n_diagnosticos_secundarios y n_procedimientos quedaron fuera pese a existir
# en el parquet: son columnas de auditoría, no features.
FEATURES_CATEGORICAS = [
    "cod_hospital",
    "sexo",
    "prevision",
    "tipo_procedencia",
    "tipo_ingreso",
    "diagnostico1_categoria",
    "procedimiento_principal",
]
FEATURES_NUMERICAS = [
    "edad", "mes_ingreso",
    "tiene_mcc", "tiene_cc", "nivel_severidad_potencial",
] + COMORBILIDADES
FEATURES = FEATURES_CATEGORICAS + FEATURES_NUMERICAS
TOP_CATEGORIAS = 200  # tope de cardinalidad para diagnostico1_categoria y procedimiento_principal

# Se entrenan DOS modelos con las mismas features, que responden preguntas distintas:
#
#   necesito_uci_uti   -> "¿va a ocupar una cama crítica?"  Pregunta de RECURSO.
#   desenlace_adverso  -> "¿le va a ir mal?"                Pregunta CLÍNICA.
#
# Hacen falta los dos porque la primera mide USO de cama, no necesidad, y el uso
# está sesgado por el racionamiento: en neumonía los mayores de 80 usan 2.6 veces
# menos UCI que los menores de 50 (8.6% vs 22.3%) mientras mueren 6 veces más
# (23.3% vs 3.9%). Un modelo entrenado solo con la primera diría "riesgo bajo"
# justo para los pacientes más frágiles.
#
# Nota: tipo_alta es FUGA como feature, pero es válida como ETIQUETA. Las etiquetas
# pueden ser post-hoc -- es lo que se quiere predecir. Solo las features tienen que
# conocerse al momento de decidir el traslado.
TARGET = "necesito_uci_uti"
TARGET_CLINICO = "desenlace_adverso"


def cargar_nombres_hospitales():
    import openpyxl

    wb = openpyxl.load_workbook(TABLAS_MAESTRAS, read_only=True, data_only=True)
    ws = wb["Hospitales"]
    nombres = {}
    for row in ws.iter_rows(values_only=True):
        cod, nombre = row[0], row[1]
        try:
            nombres[int(cod)] = nombre
        except (TypeError, ValueError):
            continue
    return nombres


def cargar_descripciones_cie10():
    """Código de categoría CIE-10 (3 caracteres, ej. 'A41') -> descripción en español."""
    import openpyxl

    wb = openpyxl.load_workbook(CIE10_DICT, read_only=True, data_only=True)
    ws = wb["CIE 10"]
    descripciones = {}
    filas = ws.iter_rows(values_only=True)
    next(filas)  # encabezado
    for _version, _codigo, _desc, categoria, *_ in filas:
        if not categoria:
            continue
        cod_cat, _, desc_cat = categoria.partition(" ")
        if cod_cat and cod_cat not in descripciones:
            descripciones[cod_cat] = desc_cat.strip().title()
    return descripciones


def entrenar_y_evaluar(train, test, objetivo, etiqueta_legible):
    """Entrena un modelo para un objetivo y devuelve todo lo que la app necesita."""
    X_train, y_train = train[FEATURES], train[objetivo]
    X_test, y_test = test[FEATURES], test[objetivo]

    modelo = HistGradientBoostingClassifier(
        categorical_features=FEATURES_CATEGORICAS,
        max_iter=200,
        learning_rate=0.08,
        max_depth=6,
        random_state=42,
    )
    print(f"\n=== {etiqueta_legible} ({objetivo}) ===")
    print("Entrenando...")
    modelo.fit(X_train, y_train)

    proba = modelo.predict_proba(X_test)[:, 1]
    auc = roc_auc_score(y_test, proba)
    ap = average_precision_score(y_test, proba)
    brier = brier_score_loss(y_test, proba)
    print(f"  ROC-AUC:        {auc:.3f}")
    print(f"  Average Prec.:  {ap:.3f}  (baseline = tasa base = {y_test.mean():.3f})")
    print(f"  Brier score:    {brier:.3f}  (menor es mejor, 0=perfecto)")

    # El caso de uso real es un comité evaluando TRASLADOS entre instituciones,
    # así que la métrica global (dominada por ingresos que no son traslado) no
    # es la que importa. Se reporta también restringida a esa subpoblación.
    es_traslado = test["tipo_procedencia"].astype(str).str.contains(
        "HOSPITAL|INSTITUCION", case=False, na=False
    )
    metricas_traslados = None
    if es_traslado.sum() > 1000 and y_test[es_traslado].nunique() > 1:
        auc_t = roc_auc_score(y_test[es_traslado], proba[es_traslado.values])
        ap_t = average_precision_score(y_test[es_traslado], proba[es_traslado.values])
        metricas_traslados = {"n": int(es_traslado.sum()), "roc_auc": auc_t,
                              "average_precision": ap_t,
                              "tasa_base": float(y_test[es_traslado].mean())}
        print(f"  -- solo traslados desde otra institución ({es_traslado.sum():,} casos) --")
        print(f"     ROC-AUC: {auc_t:.3f} | AP: {ap_t:.3f} | tasa base: {y_test[es_traslado].mean():.3f}")

    # umbrales del semáforo calibrados con percentiles reales del score (no fijos a mano):
    # rojo = 10% de mayor riesgo visto en 2024, amarillo = siguiente 20%
    umbrales = {"rojo": float(np.percentile(proba, 90)),
                "amarillo": float(np.percentile(proba, 70))}
    print(f"  Umbrales -> rojo: {umbrales['rojo']:.3f} | amarillo: {umbrales['amarillo']:.3f}")

    fpr, tpr, _ = roc_curve(y_test, proba)
    # sub-muestreo de la curva ROC para no guardar cientos de miles de puntos
    idx_roc = np.linspace(0, len(fpr) - 1, min(300, len(fpr))).astype(int)
    frac_pos, mean_pred = calibration_curve(y_test, proba, n_bins=10, strategy="quantile")

    print("  Calculando importancia de variables (muestra de 40,000 filas)...")
    rng = np.random.RandomState(42)
    muestra_idx = rng.choice(len(X_test), size=min(40000, len(X_test)), replace=False)
    pi = permutation_importance(
        modelo, X_test.iloc[muestra_idx], y_test.iloc[muestra_idx], scoring="roc_auc",
        n_repeats=5, random_state=42, n_jobs=-1,
    )
    importancia = sorted(
        zip(FEATURES, pi.importances_mean.tolist()), key=lambda x: x[1], reverse=True
    )

    tasas_historicas = {
        col: train.groupby(col, observed=True)[objetivo].mean().to_dict()
        for col in FEATURES_CATEGORICAS + COMORBILIDADES
        + ["tiene_mcc", "tiene_cc", "nivel_severidad_potencial"]
    }

    return {
        "modelo": modelo,
        "etiqueta": etiqueta_legible,
        "metricas_test_2024": {"roc_auc": auc, "average_precision": ap, "brier": brier},
        "metricas_traslados_2024": metricas_traslados,
        "tasa_base_2024": float(y_test.mean()),
        "umbrales_semaforo": umbrales,
        "curva_roc": {"fpr": fpr[idx_roc].tolist(), "tpr": tpr[idx_roc].tolist()},
        "curva_calibracion": {"mean_pred": mean_pred.tolist(), "frac_pos": frac_pos.tolist()},
        "importancia_variables": importancia,
        "tasas_historicas": tasas_historicas,
    }


def main():
    print("Cargando datos...")
    # columns= explícito: además de ahorrar memoria, impide que una columna de
    # auditoría (especialidad_medica, severidad_grd...) se cuele por accidente
    # como feature. tipo_alta entra solo para construir la etiqueta clínica.
    df = pd.read_parquet(
        DATA, columns=["anio_archivo", "tipo_actividad", "tipo_alta"] + FEATURES + [TARGET]
    )
    df["cod_hospital"] = df["cod_hospital"].astype(str)

    df[TARGET_CLINICO] = (
        (df[TARGET] == 1) | (df["tipo_alta"] == "FALLECIDO")
    ).astype(int)
    df = df.drop(columns=["tipo_alta"])

    # La población objetivo son hospitalizaciones reales sujetas a una decisión de
    # traslado a cama crítica. Se sacan dos grupos que no lo son:
    #
    # 1) Cirugía mayor ambulatoria: opera y se va el mismo día, 1.2% de tasa
    #    UCI/UTI. Solo inflaba la métrica global con negativos fáciles
    #    (ROC-AUC 0.931 incluyéndola vs 0.914 sobre la población real).
    # 2) Sesiones ambulatorias de quimioterapia (dx Z51) y diálisis (dx Z49):
    #    el paciente viene a una prestación puntual y se va.
    #
    # El corte de (2) es por DIAGNÓSTICO principal, no por procedimiento: el
    # procedimiento 39.95 (hemodiálisis) tiene 19.4% de tasa UCI porque también se
    # dializa a pacientes críticos con falla renal aguda, y excluirlo borraría
    # pacientes reales. TIPO_ACTIVIDAD tampoco sirve para esto: la categoría
    # 'HOSPITALIZACIÓN DIURNA' solo existe en 2019.
    n_antes = len(df)
    es_ambulatorio = (
        df["tipo_actividad"].astype(str).str.contains("CMA", na=False)
        | df["diagnostico1_categoria"].isin(["Z49", "Z51"])
    )
    df = df[~es_ambulatorio]
    print(f"Excluidas {n_antes - len(df):,} atenciones ambulatorias (CMA, quimio, diálisis) "
          f"de {n_antes:,} admisiones.")
    df = df.drop(columns=["tipo_actividad"])

    train_full = df[df["anio_archivo"].isin(ANIOS_TRAIN)]

    stats_hosp = train_full.groupby("cod_hospital").agg(
        n=("cod_hospital", "size"), tasa=(TARGET, "mean")
    )
    hospitales_validos = stats_hosp[
        (stats_hosp["n"] >= MIN_CASOS_HOSPITAL) & (stats_hosp["tasa"] >= MIN_TASA_HOSPITAL)
    ].index
    excluidos = stats_hosp.index.difference(hospitales_validos)
    print(
        f"Hospitales válidos: {len(hospitales_validos)} / {len(stats_hosp)} "
        f"(excluidos {len(excluidos)} por bajo volumen o sin UCI/UTI propia)"
    )

    df = df[df["cod_hospital"].isin(hospitales_validos)].copy()
    train = df[df["anio_archivo"].isin(ANIOS_TRAIN)].copy()
    test = df[df["anio_archivo"] == ANIO_TEST].copy()
    print(f"Train: {len(train):,} filas ({min(ANIOS_TRAIN)}-{max(ANIOS_TRAIN)}) | "
          f"Test: {len(test):,} filas ({ANIO_TEST})")

    # HistGradientBoosting exige cardinalidad <= 255 por feature categórica;
    # diagnostico1_categoria y procedimiento_principal tienen miles de valores ->
    # nos quedamos con los más frecuentes (calculado solo con train, sin fuga)
    # y el resto va a "OTRO"
    top_categorias_diag = train["diagnostico1_categoria"].value_counts().head(TOP_CATEGORIAS).index
    for d in (train, test):
        d["diagnostico1_categoria"] = d["diagnostico1_categoria"].where(
            d["diagnostico1_categoria"].isin(top_categorias_diag), "OTRO"
        )
    top_procedimientos = train["procedimiento_principal"].value_counts().head(TOP_CATEGORIAS).index
    for d in (train, test):
        d["procedimiento_principal"] = d["procedimiento_principal"].where(
            d["procedimiento_principal"].isin(top_procedimientos)
            | (d["procedimiento_principal"] == "SIN_PROCEDIMIENTO"),
            "OTRO",
        )

    for col in FEATURES_CATEGORICAS:
        train[col] = train[col].astype("category")
        test[col] = pd.Categorical(test[col], categories=train[col].cat.categories)

    print(f"Tasa base -> cama crítica: {train[TARGET].mean():.3f} | "
          f"desenlace adverso: {train[TARGET_CLINICO].mean():.3f}")

    cama = entrenar_y_evaluar(train, test, TARGET, "Probabilidad de cama crítica")
    clinico = entrenar_y_evaluar(train, test, TARGET_CLINICO, "Riesgo de desenlace adverso")

    # metadata para poblar los selects de la app. Se construye desde TRAIN, no
    # desde df completo: si saliera de df, la app ofrecería valores que solo
    # aparecen en 2024 y que el modelo nunca vio (los mandaba a NaN en silencio).
    opciones = {
        col: sorted(train[col].cat.categories.astype(str).tolist())
        for col in FEATURES_CATEGORICAS
    }
    nombres_hospitales = cargar_nombres_hospitales()
    opciones["cod_hospital_nombre"] = {
        cod: nombres_hospitales.get(int(cod), f"Hospital {cod}") for cod in opciones["cod_hospital"]
    }
    descripciones_cie10 = cargar_descripciones_cie10()
    opciones["diagnostico1_descripcion"] = {
        cod: descripciones_cie10.get(cod, "Sin descripción disponible")
        for cod in opciones["diagnostico1_categoria"]
    }
    opciones["diagnostico1_descripcion"]["OTRO"] = "Otro diagnóstico (fuera de los 200 más frecuentes)"

    # Las claves de 'cama' quedan también en la raíz por compatibilidad con el
    # código de la app que ya las usaba (modelo, umbrales_semaforo, etc.).
    joblib.dump(
        {
            **cama,
            "features_categoricas": FEATURES_CATEGORICAS,
            "features_numericas": FEATURES_NUMERICAS,
            "opciones": opciones,
            "cama_critica": cama,
            "riesgo_clinico": clinico,
        },
        MODELO_OUT,
    )
    print(f"\nModelo guardado en {MODELO_OUT}")


if __name__ == "__main__":
    main()
